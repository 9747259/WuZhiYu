from time import sleep
import time

from openpyxl.worksheet.filters import FilterColumn
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import re

option = webdriver.ChromeOptions()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

chrome_path = executable_path = 'C:/Users/Administrator/AppData/Local/Google/Chrome/Application/chromedriver.exe'
driver = webdriver.Chrome(chrome_path, options=option)


def code_chioce():
    for y in range(1, length + 1):
        order_number = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                                    '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[6]/div/span' % y).text
        code_number = driver.find_element_by_xpath(
            '/html/body/div[1]/section/section/section/main/div/div[1]/div[2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span' % y).text

        for x in range(2, max_row + 1):
            if str(mysheet.cell(x, 2).value) == order_number:
                if str(mysheet.cell(x, 3).value) == code_number:
                    driver.find_element_by_xpath(
                        '/html/body/div[1]/section/section/section/main/div/div[1]/div[2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[1]/div/span/span' % y).click()
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div[2]/div/div/div/div/div[1]/div/div[1]/div/button[1]').click()

def code_execute():
    length2 = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr'))
    for y in range(1, length2+1):
        final_order = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[5]/div/span' % y).text
        final_code = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span' % y).text
        final_number = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[9]/div/span' % y).text

        for x in range(2, max_row+1):
            if str(mysheet.cell(x, 2).value) == final_order:
                if str(mysheet.cell(x, 3).value) == final_code:
                    if mysheet.cell(x, 7).value == float(final_number):
                        '''
                        path_action = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                     '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                     '2]/table/tbody/tr[%d]/td[14]/div/div/div/div'%y)
                        action = ActionChains(driver).click(path_action).send_keys(str(mysheet.cell(x, 1).value)).pause(3).move_by_offset(2,4).click().perform()
                        '''

                        '''
                        #点击下拉菜单，并选择库位
                        driver.find_element_by_xpath(
                            '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                            '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td['
                            '14]/div/div/div/span' % y).click()
                        sleep(2)
                        
                        location_number = select_location(mysheet.cell(x, 1).value)
                        driver.find_element_by_xpath('/html/body/div[last()]/div/div/div/ul/li[%d]' % location_number).click()
                        '''
                        driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                     '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                     '2]/table/tbody/tr[%d]/td['
                                                     '1]/div/span/span' % y).click()
                        mysheet.cell(x, 9).value = '物资域已收到货'
    mybook.save(r'D:\工作相关\到库收货.xlsx')

def select_location(temp_location):
    if temp_location == '0001':
        return 1
    elif temp_location == '0005':
        return 5
    elif temp_location == '0004':
        return 4
    elif temp_location == '2006':
        return 52
    elif temp_location == '2007':
        return 53
    elif temp_location == '2008':
        return 54
    elif temp_location == '4002':
        return 70
    else:
        return 3


if __name__ == "__main__":
    arr = []
    temp_number = 0
    b = 0

    el_frame = driver.find_element_by_xpath("/html/body/form/div[3]/div[3]/div[2]/div[6]/div/iframe")
    driver.switch_to.frame(el_frame)  # 非常重要的一步，跳转到内部的iframe

    length = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                               '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr'))

    mybook = openpyxl.load_workbook(r'D:\工作相关\到库收货.xlsx')
    mysheet = mybook.active
    max_row = mysheet.max_row

    code_chioce()
    sleep(5)
    code_execute()
    print("已完成")



