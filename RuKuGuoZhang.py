from time import sleep

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import PatternFill
import openpyxl


option = webdriver.ChromeOptions()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

chrome_path = executable_path = 'C:/Users/Administrator/AppData/Local/Google/Chrome/Application/chromedriver.exe'
driver = webdriver.Chrome(chrome_path, options=option)

def table_head(row):
    '''输入订单号'''
    driver.find_element_by_xpath(
        '/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div[2]/div[1]/div['
        '1]/div/div/div/div[1]/div/div[8]/div/div[2]/div/div/div/div/ul/li/div/span[1]/span/span[1]/input').send_keys(
        mysheet.cell(row, 2).value)

    '''点击查询按钮'''
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div['
                                 '2]/div[1]/div[1]/div/div/div/div[2]/div/div/button[1]').click()


def order_deal():
    global arr, temp_number, b
    temp_number = 0
    flag_order_deal = 0
    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    sleep(1)
    length = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div['
                                               '3]/div[1]/div[2]/div[1]/div[2]/div/div/div/div/div[3]/div[2]/div['
                                               '1]/div[2]/table/tbody/tr'))
    if length == 0:
        '''点击订单取消按钮'''
        driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div['
                                     '2]/div[1]/div[1]/div/div/div/div[1]/div/div[8]/div/div['
                                     '2]/div/div/div/div/ul/li/div/span[1]/span/span[1]/span').click()
        sleep(2)
        b = 0
    else:
        for sheet_row in arr:
            for system_row in range(1, length + 1):
                if mysheet.cell(sheet_row, 3).value == driver.find_element_by_xpath(
                        '/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div[2]/div[1]/div['
                        '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span' %
                        system_row).text:
                    if mysheet.cell(sheet_row, 7).value == float(driver.find_element_by_xpath('/html/body/div['
                                                                                              '1]/section/section'
                                                                                              '/section/main/div/div'
                                                                                              '/div/div[3]/div['
                                                                                              '1]/div[2]/div[1]/div['
                                                                                              '2]/div/div/div/div'
                                                                                              '/div[3]/div[2]/div['
                                                                                              '1]/div['
                                                                                              '2]/table/tbody/tr['
                                                                                              '%d]/td[6]/div/span' %
                                                                                              system_row).text):
                        driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div['
                                                     '3]/div[1]/div[2]/div[1]/div[2]/div/div/div/div/div[3]/div['
                                                     '2]/div[1]/div[2]/table/tbody/tr[%d]/td[1]/div/span/span' %
                                                     system_row).click()
                        mysheet.cell(sheet_row, 9).value = '已过账'
                        #mysheet.cell(sheet_row, 7).fill = orange_fill
                        flag_order_deal = 1

                        break

        if flag_order_deal == 1:
            sleep(2)
            driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div['
                                     '2]/div[1]/div[2]/div/div/div/div/div[1]/div/div[1]/div/button[1]').click()
            sleep(2)
            driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/div/div[2]/div[3]/div/button[2]').click()
            flag_order_deal = 0
        driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div['
                                     '2]/div[1]/div[1]/div/div/div/div[1]/div/div[8]/div/div['
                                     '2]/div/div/div/div/ul/li/div/span[1]/span/span[1]/span').click()
        mybook.save(r'D:\工作相关\入库过账.xlsx')

if __name__ == "__main__":
    arr = []
    temp_number = 0
    b = 0
    sheet_start_number = int(input("请输入表格开始行号:"))

    el_frame = driver.find_element_by_xpath("/html/body/form/div[3]/div[3]/div[2]/div[6]/div/iframe")
    driver.switch_to.frame(el_frame)  # 非常重要的一步，跳转到内部的iframe

    mybook = openpyxl.load_workbook(r'D:\工作相关\入库过账.xlsx')
    mysheet = mybook.active
    max_row = mysheet.max_row

    for row in range(sheet_start_number, max_row):
        if mysheet.cell(row + 1, 2).value == mysheet.cell(row, 2).value:
            if arr == []:
                arr = [row, row + 1]
            else:
                arr.append(row + 1)
        else:
            if arr == []:
                arr.append(row)
                sleep(1)  # 非常重要，等网页读取数据完成，不然影响到后面数据选择
                table_head(row)
                order_deal()
                arr = []
            else:
                table_head(row)
                sleep(1)  # 非常重要，等网页读取数据完成，不然影响到后面数据选择
                order_deal()
                arr = []

        print('需收料表总共%d行,当前行号是：%d,还剩%d行未完成。' % (max_row,row+1,max_row-row+1))
    print("已完成！！！")