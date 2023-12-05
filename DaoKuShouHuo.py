from time import sleep

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.remote import switch_to
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import PatternFill
import openpyxl
import re

option = webdriver.ChromeOptions()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

chrome_path = executable_path = 'C:/Users/Administrator/AppData/Local/Google/Chrome/Application/chromedriver.exe'
driver = webdriver.Chrome(chrome_path, options=option)


def table_head(row):
    '''打开折叠'''
    driver.find_element_by_xpath(
        '/html/body/div/section/section/section/main/div/div[1]/div[1]/div/div/div/div[2]/div/div/span').click()
    '''输入订单号'''
    driver.find_element_by_xpath(
        '/html/body/div[1]/section/section/section/main/div/div[1]/div[1]/div/div/div/div[1]/div/div[6]/div/div['
        '2]/div/div/div/div/ul/li/div/span[1]/span/span[1]/input').send_keys(
        mysheet.cell(row, 2).value)

    '''点击查询按钮'''
    driver.find_element_by_xpath('//*[@id="search-and-tools"]/div/div/div/div[2]/div/div/button[1]').click()


def order_deal():
    global arr, arr2, temp_number, b

    temp_number = 0
    length = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                      '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr'))
    # 检查该订单是否有物料可以收料
    if length == 0:
        '''点击订单取消按钮'''
        cancel_button()
        b = 0
    else:
        for sheet_row in arr:
            for system_row in range(1, length + 1):
                if mysheet.cell(sheet_row + 1, 3).value == mysheet.cell(sheet_row, 3).value:
                    if arr2 == []:
                        arr2 = [sheet_row, sheet_row+1]
                    else:
                        arr2.append(sheet_row)
                else:
                    if arr2 == []:
                        if mysheet.cell(sheet_row, 3).value == driver.find_element_by_xpath(
                                '/html/body/div/section/section/section/main/div/div[1]/div[2]/div/div/div/div/div[3]/div['
                                '2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div'%system_row).text:
                            if mysheet.cell(sheet_row, 7).value <= float(driver.find_element_by_xpath('/html/body/div['
                                                                                                      '1]/section/section'
                                                                                                      '/section/main/div/div['
                                                                                                      '1]/div['
                                                                                                      '2]/div/div/div/div'
                                                                                                      '/div[3]/div[2]/div['
                                                                                                      '1]/div['
                                                                                                      '2]/table/tbody/tr['
                                                                                                      '%d]/td['
                                                                                                      '7]/div'%system_row).text):
                                driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                                             '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr['
                                                             '%d]/td[1]/div/span'%system_row).click()
                                b = 1
                                break
                    else:
                        for x in arr2:
                            temp_number = mysheet.cell(x, 7).value + temp_number

                        if mysheet.cell(sheet_row, 3).value == driver.find_element_by_xpath(
                                '/html/body/div/section/section/section/main/div/div[1]/div[2]/div/div/div/div/div[3]/div['
                                '2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div'%system_row).text:
                            if temp_number <= float(driver.find_element_by_xpath('/html/body/div['
                                                                                                      '1]/section/section'
                                                                                                      '/section/main/div/div['
                                                                                                      '1]/div['
                                                                                                      '2]/div/div/div/div'
                                                                                                      '/div[3]/div[2]/div['
                                                                                                      '1]/div['
                                                                                                      '2]/table/tbody/tr['
                                                                                                      '%d]/td['
                                                                                                      '7]/div' % system_row).text):
                                driver.find_element_by_xpath(
                                    '/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                    '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr['
                                    '%d]/td[1]/div/span' % system_row).click()
                                b = 1
                                break
                        arr2 = []
                        temp_number = 0
        ''''''
        if b == 0:
            cancel_button()

def confirm():
    global arr
    arr3 = []
    final_temp_number = 0
    flag_confirm = 0

    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                 '2]/div/div/div/div/div[1]/div/div[1]/div/button[1]').click()
    length2 = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr'))
    sleep(1)
    for final_sheet_row in arr:
        flag_y = 0
        if mysheet.cell(final_sheet_row + 1, 3).value == mysheet.cell(final_sheet_row, 3).value:
            if arr3 == []:
                arr3 = [final_sheet_row, final_sheet_row + 1]
            else:
                arr3.append(final_sheet_row)
        else:
            for final_system_row in range(1, length2+1):
                if arr3 == []:
                    if mysheet.cell(final_sheet_row, 3).value == driver.find_element_by_xpath(
                            '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                            '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span' % final_system_row).text:
                        if mysheet.cell(final_sheet_row, 7).value <= float(driver.find_element_by_xpath('/html/body'
                                                                                                        '/div['
                                                                                                        '1]/section/section/section/main/div/div[2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[9]/div/span' % final_system_row).text):
                            path_action = driver.find_element_by_xpath(
                                '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td['
                                '10]/div/span/div/div/div[2]/input' % final_system_row)

                            if driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                            '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                            '2]/table/tbody/tr[%d]/td[13]/div/span'%final_system_row).text == '上架':
                                driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                             '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                             '2]/table/tbody/tr[%d]/td[13]/div/span'%final_system_row).click()
                                sleep(2)
                                if final_system_row == 1:
                                    driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/div/div[2]/div[2]/div/div['
                                                                 '3]/div[1]/div/div/div/div/div/div/div/div['
                                                                 '2]/table/tbody/tr[1]/td[1]/span/label/span/input').click()
                                    driver.find_element_by_xpath(
                                        '/html/body/div[4]/div/div[2]/div/div[2]/div[3]/div/button[2]').click()
                                else:
                                    driver.find_element_by_xpath(
                                        '/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div/div['
                                        '3]/div[1]/div/div/div/div/div/div/div/div['
                                        '2]/table/tbody/tr[1]/td[1]/span/label/span/input').click()
                                    driver.find_element_by_xpath(
                                        '/html/body/div[5]/div/div[2]/div/div[2]/div[3]/div/button[2]').click()

                                sleep(2)
                                driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                             '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                             '2]/table/tbody/tr[%d]/td['
                                                             '1]/div/span/span' % final_system_row).click()
                                flag_confirm = 1
                                sleep(2)
                            '''点击下拉菜单，并选择库位'''
                            driver.find_element_by_xpath(
                                '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td['
                                '14]/div/div/div/span' % final_system_row).click()
                            sleep(2)
                            location_number = select_location(mysheet.cell(final_sheet_row, 1).value)
                            if flag_confirm == 0:
                                driver.find_element_by_xpath(
                                    '/html/body/div[%d]/div/div/div/ul/li[%d]' % (final_system_row+3,location_number)).click()
                            else:
                                driver.find_element_by_xpath(
                                    '/html/body/div[%d]/div/div/div/ul/li[%d]' % (final_system_row+4,location_number)).click()
                                flag_confirm = 0
                            '''输入数量'''
                            ActionChains(driver).double_click(path_action).send_keys(
                                str(mysheet.cell(final_sheet_row, 7).value)).perform()
                            driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                         '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                         '2]/table/tbody/tr[%d]/td['
                                                         '1]/div/span/span' % final_system_row).click()
                            '''保存EXCEL表数据'''
                            mysheet.cell(final_sheet_row, 9).value = '物资域已收到货'
                            break
                else:
                    if flag_y == 0:
                        for y in arr3:
                            final_temp_number = mysheet.cell(y, 7).value + final_temp_number
                        flag_y = 1
                    if mysheet.cell(final_sheet_row, 3).value == driver.find_element_by_xpath(
                            '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                            '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span'
                            %final_system_row).text:
                        if final_temp_number <= float(driver.find_element_by_xpath(
                                '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[9]/div/span' %
                                final_system_row).text):
                            path_action = driver.find_element_by_xpath(
                                '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td['
                                '10]/div/span/div/div/div[2]/input' % final_system_row)
                            if driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                         '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                         '2]/table/tbody/tr/td[13]/div/span').text == '上架':
                                driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                             '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                             '2]/table/tbody/tr/td[13]/div/span').click()
                                sleep(2)
                                driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/div/div[2]/div[2]/div/div['
                                                             '3]/div[1]/div/div/div/div/div/div/div/div['
                                                             '2]/table/tbody/tr[1]/td[1]/span/label/span/input').click()
                                sleep(1)
                                driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/div/div[2]/div[3]/div/button[2]').click()
                                sleep(2)
                                driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                             '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                             '2]/table/tbody/tr[%d]/td['
                                                             '1]/div/span/span' % final_system_row).click()

                                flag_confirm = 1
                                sleep(2)
                            '''点击下拉菜单，并选择库位'''
                            driver.find_element_by_xpath(
                                '/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                '1]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td['
                                '14]/div/div/div/span' % final_system_row).click()
                            sleep(2)
                            location_number = select_location(mysheet.cell(final_sheet_row, 1).value)
                            if flag_confirm == 0:
                                driver.find_element_by_xpath(
                                    '/html/body/div[%d]/div/div/div/ul/li[%d]' % (final_system_row + 3, location_number)).click()
                            else:
                                driver.find_element_by_xpath(
                                    '/html/body/div[%d]/div/div/div/ul/li[%d]' % (final_system_row + 4, location_number)).click()
                                flag_confirm = 0
                            '''输入数量'''
                            ActionChains(driver).double_click(path_action).send_keys(str(final_temp_number)).perform()
                            driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div['
                                                         '2]/div/div[1]/div/div/div/div/div[3]/div[2]/div[1]/div['
                                                         '2]/table/tbody/tr[%d]/td['
                                                         '1]/div/span/span'%final_system_row).click()
                            '''保存EXCEL表数据'''
                            mysheet.cell(final_sheet_row, 9).value = '物资域已收到货'
                            arr3 = []
                            final_temp_number = 0
                            break

    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                 '1]/div/div/div/div/div[1]/div/div[1]/div/button[1]') .click()
    sleep(2)
    try:
        driver.find_element_by_css_selector('body > div:nth-child(%d) > div > '
                                            'div.ant-modal-wrap.ant-modal-centered.ant-modal-confirm-centered > div > '
                                            'div.ant-modal-content > div > div > div.ant-modal-confirm-btns > '
                                            'button.ant-btn.ant-btn-primary'%(final_system_row+10)).click()
        sleep(2)
        driver.find_element_by_css_selector(
            'body > div:nth-child(9) > div > div.ant-modal-wrap.ant-modal-centered > div > div.ant-modal-content > '
            'div.ant-modal-footer > button:nth-child(1)').click()
    except:
        driver.find_element_by_css_selector('body > div:nth-child(12) > div > '
                                            'div.ant-modal-wrap.ant-modal-centered.ant-modal-confirm-centered > div > '
                                            'div.ant-modal-content > div > div > div.ant-modal-confirm-btns > '
                                            'button.ant-btn.ant-btn-primary').click()
        sleep(2)
        driver.find_element_by_css_selector(
            'body > div:nth-child(10) > div > div.ant-modal-wrap.ant-modal-centered > div > div.ant-modal-content > '
            'div.ant-modal-footer > button:nth-child(1)').click()
        flag_confirm = 0
    '''保存EXCEL表的数据'''
    mybook.save(r'D:\工作相关\到库收货.xlsx')
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[2]/div/div['
                                 '1]/div/div/div/div/div[1]/div/div[1]/div/button[3]') .click()
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div[1]/div/div/div/div['
                                 '1]/div/div[6]/div/div[2]/div/div/div/div/ul/li/div/span[1]/span/span['
                                 '1]/span').click()
    driver.find_element_by_xpath(
        '/html/body/div/section/section/section/main/div/div[1]/div[1]/div/div/div/div[2]/div/div/span').click()

def select_location(temp_location):
    if temp_location == '0001':
        return 1
    elif temp_location == '0005':
        return 5
    elif temp_location == '0004':
        return 4
    elif temp_location == '2006':
        return 52
    elif temp_location == '2007':
        return 53
    elif temp_location == '2008':
        return 54
    elif temp_location == '4002':
        return 70
    else:
        return 3


def cancel_button():
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                 '1]/div/div/div/div[1]/div/div[6]/div/div[2]/div/div/div/div/ul/li/div/span['
                                 '1]/span/span[1]/span').click()
    sleep(1)
    driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div[1]/div['
                                 '1]/div/div/div/div[2]/div/div/span/i').click()

def confirm_button(final_sheet_row, final_system_row):
    pass


if __name__ == "__main__":
    arr = []
    arr2 = []
    temp_number = 0
    confirm_ok = 1
    b = 0
    el_frame = driver.find_element_by_xpath("/html/body/form/div[3]/div[3]/div[2]/div[6]/div/iframe")
    driver.switch_to.frame(el_frame)  # 非常重要的一步，跳转到内部的iframe

    mybook = openpyxl.load_workbook(r'D:\工作相关\到库收货.xlsx')
    mysheet = mybook.active
    max_row = mysheet.max_row
    startrow = int(input('请输入表格开始行号：'))

    '''循环EXCEL表的每一行'''
    for row in range(startrow, max_row):
        if mysheet.cell(row + 1, 2).value == mysheet.cell(row, 2).value:
            if arr == []:
                arr = [row, row+1]
            else:
                arr.append(row+1)
        else:
            if arr == []:
                arr.append(row)
                sleep(2)  #非常重要，等网页读取数据完成，不然影响到后面数据选择
                '''网页的表头录入程序'''
                table_head(row)
                '''网页的物料选择程序'''
                order_deal()
                if confirm_ok == 0:
                    confirm_ok = 1
                    break
                sleep(2)
                if b == 1:
                    '''网页的物料数量录入程序，如果没有可以录入的，就不执行'''
                    confirm()
                    b = 0
                arr = []
            else:
                table_head(row)
                sleep(2)  #非常重要，等网页读取数据完成，不然影响到后面数据选择
                order_deal()
                if confirm_ok == 0:
                    confirm_ok = 1
                    break
                sleep(2)
                if b == 1:
                    confirm()
                    b = 0
                arr = []
        print("总行号是：%d,当前的行号是:%d,还剩%d条未完成" % (max_row, row+1, (max_row-row+1)))
    print("已完成")