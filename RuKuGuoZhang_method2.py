from time import sleep

from openpyxl.worksheet.filters import FilterColumn
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import re

option = webdriver.ChromeOptions()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

chrome_path = executable_path = 'C:/Users/Administrator/AppData/Local/Google/Chrome/Application/chromedriver.exe'
driver = webdriver.Chrome(chrome_path, options=option)

def code_chioce(length):
    for y in range(1, length + 1):
        order_number = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div['
                                                    '3]/div[1]/div[2]/div[1]/div[2]/div/div/div/div/div[3]/div['
                                                    '2]/div[1]/div[2]/table/tbody/tr[%d]/td[9]/div' % y).text

        code_number = driver.find_element_by_xpath(
            '/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div[2]/div[1]/div['
            '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[3]/div/span' % y).text
        total_number = driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div[2]/div[1]/div[2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[6]/div/span'%y).text
        for x in range(2, max_row + 1):
            if str(mysheet.cell(x, 2).value) == order_number:
                if str(mysheet.cell(x, 3).value) == code_number:
                    #if str(mysheet.cell(x, 7).value) == str(total_number):
                    driver.find_element_by_xpath(
                        '/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div[2]/div[1]/div['
                        '2]/div/div/div/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr[%d]/td[1]/div/span/span' %
                        y).click()
                    mysheet.cell(x, 9).value = '已过账'
                    mybook.save(r'D:\工作相关\入库过账.xlsx')
                    break


    #driver.find_element_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div[3]/div[1]/div['
    #                            '2]/div[1]/div[2]/div/div/div/div/div[1]/div/div[1]/div/button[1]').click()


def code_execute():
    pass


if __name__ == "__main__":
    arr = []
    temp_number = 0
    b = 0

    el_frame = driver.find_element_by_xpath("/html/body/form/div[3]/div[3]/div[2]/div[6]/div/iframe")
    driver.switch_to.frame(el_frame)  # 非常重要的一步，跳转到内部的iframe

    length = llength = len(driver.find_elements_by_xpath('/html/body/div[1]/section/section/section/main/div/div/div/div['
                                               '3]/div[1]/div[2]/div[1]/div[2]/div/div/div/div/div[3]/div[2]/div['
                                               '1]/div[2]/table/tbody/tr'))

    mybook = openpyxl.load_workbook(r'D:\工作相关\入库过账.xlsx')
    mysheet = mybook.active
    max_row = mysheet.max_row

    code_chioce(length)
    print("已完成")



